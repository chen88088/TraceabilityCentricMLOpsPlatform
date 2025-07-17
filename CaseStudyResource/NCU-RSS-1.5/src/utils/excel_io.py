"""Module importing the excel file"""
from openpyxl import load_workbook


class ExcelInfo:
    """Module importing the excel file"""

    def __init__(self, excel_path, round_number, number_of_frames):
        """
        self.training_and_validation_ds_frame_codename_list:
            紀錄使用於training和validation的frame codename
            e.g. [0, 2, 4, 7]

        self.testing_ds_frame_codename_list:
            紀錄使用於testing的frame codename
            e.g. [1, 3, 5]


        """
        self.round_number = round_number
        self.training_and_validation_ds_frame_codename_list = []
        self.testing_ds_frame_codename_list = []
        self.number_of_frames = number_of_frames
        self.codename_table = self.read_codename_table(excel_path)
        self.read_ds_frame_number_list_in_excel_workbook(excel_path)

    def read_ds_frame_number_list_in_excel_workbook(self, excel_path):
        """
        遍歷Excel中, 一個Round使用的frame資料
        """
        loaded_wb = load_workbook(excel_path)

        sheet_name = f"R{str(self.round_number)}"
        sheet = loaded_wb.get_sheet_by_name(sheet_name)  # 紀錄該round資料的sheet

        # 忽略欄位名稱(excel第1列)，遍歷第一筆資料(excel第2列)到最後一筆資料
        # max_index_row = sheet.max_row-1 # sheet.max_row由1開始計算,但index由0開始計算

        # sheet的第一列資料，row_of_sheet為1 => 第一筆~最後一筆
        # index_row = 1 對應excel第二列(忽略欄位名稱) 1~max_index_row
        for row_of_sheet in range(3, self.number_of_frames * 2 + 2 + 1):
            print(
                f"{row_of_sheet} C:{sheet['c' + str(row_of_sheet)].value} D:{sheet['d' + str(row_of_sheet)].value}")

            if sheet['c' + str(row_of_sheet)].value is not None:
                codename = self.frame_number_and_shot_date_to_codename(
                    sheet['c' + str(row_of_sheet)].value)
                self.testing_ds_frame_codename_list.append(codename)

            if sheet['d' + str(row_of_sheet)].value is not None:
                codename = self.frame_number_and_shot_date_to_codename(
                    sheet['d' + str(row_of_sheet)].value)
                self.training_and_validation_ds_frame_codename_list.append(
                    codename)

    def read_codename_table(self, excel_path):
        """
        以dictionary紀錄frame_codename相對應的frame_number_and_shot_date的table

        e.g.
        codename_table = {
            "0":"94191004_181006z",
            "1":"94191006_181006z",
            ...
        }
        """
        loaded_wb = load_workbook(excel_path)
        sheet_name = "codename table"
        sheet = loaded_wb.get_sheet_by_name(sheet_name)
        codename_table = {}
        # 1st +1 is for 2 start, 2nd +1 is for exclusive right bound
        for row_of_sheet in range(2, self.number_of_frames + 1 + 1):
            codename_table[sheet['b' + str(row_of_sheet)
                                 ].value] = sheet['a' + str(row_of_sheet)].value

        return codename_table

    def frame_number_and_shot_date_to_codename(
            self, frame_number_and_shot_date):
        for key, value in self.codename_table.items():
            if value == frame_number_and_shot_date:
                return key
